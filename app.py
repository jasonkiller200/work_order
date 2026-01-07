from flask import Flask, render_template, jsonify, redirect, url_for, request, session, flash, send_file, make_response
import os
import pandas as pd
import threading
import time
import logging
import math # 導入 math 模組來檢查 NaN
from datetime import datetime # 導入 datetime 模組
import pytz # 導入 pytz 模組
import json # 導入 json 模組
import io # 導入 io 模組
from urllib.parse import quote # 導入 quote 函式
from consolidate_specifications import consolidate_spec_files
from waitress import serve # 導入 waitress 的 serve 函式
import xlrd # 導入 xlrd 函式庫 (用於 .xls 檔案)
from app.config.settings import Config # 導入應用程式設定

# --- 配置日誌 ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app_errors.log"),
        logging.StreamHandler()
    ]
)
app_logger = logging.getLogger(__name__)

# --- 全域變數 ---
data_cache = {"A": None, "B": None}
live_cache_pointer = "A"
cache_lock = threading.Lock()

# --- 訂單備註與版本快取 ---
order_note_cache = {}
order_note_cache_lock = threading.Lock()

def load_order_notes_to_cache():
    app_logger.info("開始載入訂單備註與版本快取...")
    file_path = r'Q:\\G003\\生產排程\\8週生產排程紀錄\\裝配進度&缺料情報\\第一廠\\組件課\\組件1-20新增缺料查詢功能.xls'
    sheet_name = '整合資料'

    temp_cache = {}
    try:
        workbook = xlrd.open_workbook(file_path, formatting_info=True)
        sheet = workbook.sheet_by_name(sheet_name)
        
        # --- 動態尋找欄位索引 ---
        header_row = sheet.row_values(0)
        order_id_col_idx = 5 # Column F is index 5 (0-indexed)
        version_col_idx = -1
        try:
            version_col_idx = header_row.index('訂單版本')
        except ValueError:
            app_logger.error(f"在 {file_path} 的 '{sheet_name}' 頁籤中找不到 '訂單版本' 欄位。將無法載入版本資訊。")
            # 即使找不到版本欄位，我們仍然繼續，只是不載入版本資訊

        # --- 遍歷資料列 ---
        for row_idx in range(1, sheet.nrows): # 從第 1 行開始，跳過標頭
            order_id_cell = sheet.cell(row_idx, order_id_col_idx)
            order_id = str(order_id_cell.value).strip()

            if order_id:
                note_text = None
                if (row_idx, order_id_col_idx) in sheet.cell_note_map:
                    note = sheet.cell_note_map[(row_idx, order_id_col_idx)]
                    note_text = note.text
                
                version_text = None
                if version_col_idx != -1:
                    version_cell = sheet.cell(row_idx, version_col_idx)
                    version_text = str(version_cell.value).strip()

                # 只有在備註或版本存在時才加入快取
                if note_text or version_text:
                    temp_cache[order_id] = {'note': note_text, 'version': version_text}

        with order_note_cache_lock:
            global order_note_cache
            order_note_cache = temp_cache
        app_logger.info(f"訂單備註與版本快取載入完成。共載入 {len(order_note_cache)} 條紀錄。")

    except FileNotFoundError:
        app_logger.error(f"載入訂單快取失敗：找不到檔案: {file_path}")
    except xlrd.biffh.XLRDError as e:
        app_logger.error(f"載入訂單快取失敗：使用 xlrd 讀取 {file_path} 時發生錯誤: {e}", exc_info=True)
    except Exception as e:
        app_logger.error(f"載入訂單快取失敗：發生未知錯誤: {e}", exc_info=True)

def update_order_note_cache_periodically():
    while True:
        time.sleep(3600) # 每小時更新一次快取 (3600 秒)
        app_logger.info("背景執行緒：準備更新訂單備註與版本快取...")
        load_order_notes_to_cache()

# --- 瀏覽次數檔案路徑 ---
VIEWS_FILE = 'page_views.json'

# --- 瀏覽次數讀寫輔助函式 ---
def read_views():
    if os.path.exists(VIEWS_FILE):
        try:
            with open(VIEWS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            app_logger.error(f"讀取 {VIEWS_FILE} 時發生 JSON 解碼錯誤: {e}", exc_info=True)
            return {}
        except Exception as e:
            app_logger.error(f"讀取 {VIEWS_FILE} 時發生錯誤: {e}", exc_info=True)
            return {}
    return {}

def write_views(views_data):
    try:
        with open(VIEWS_FILE, 'w', encoding='utf-8') as f:
            json.dump(views_data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        app_logger.error(f"寫入 {VIEWS_FILE} 時發生錯誤: {e}", exc_info=True)

# --- 記錄頁面訪問次數和 IP ---
def record_page_view(page_name, ip_address):
    views_data = read_views()
    
    if page_name not in views_data:
        views_data[page_name] = {
            "total_views": 0,
            "ip_access_times": {}
        }
    
    views_data[page_name]["total_views"] += 1
    
    if ip_address not in views_data[page_name]["ip_access_times"]:
        views_data[page_name]["ip_access_times"][ip_address] = []
    
    # 獲取台灣時間
    taiwan_tz = pytz.timezone('Asia/Taipei')
    current_time = datetime.now(taiwan_tz).isoformat()
    
    views_data[page_name]["ip_access_times"][ip_address].append(current_time)
    
    write_views(views_data)

# --- 資料處理核心邏輯 ---
def load_and_process_data():
    app_logger.info("開始載入與處理資料...")
    try:
        df_inventory = pd.read_excel(r'P:\\F004\\MPS維護\\零件庫存.XLSX')
        df_wip_parts = pd.read_excel(r'P:\F004\MPS維護\撥料.XLSX')
        df_finished_parts = pd.read_excel(r'P:\F004\MPS維護\成品撥料.XLSX')
        df_prep_semi_finished = pd.read_excel(r'P:\F004\MPS維護\預備半品用料.xlsx')

        # 根據訂單號碼首位數字篩選
        df_wip_parts['訂單'] = df_wip_parts['訂單'].astype(str)
        df_wip_parts = df_wip_parts[df_wip_parts['訂單'].str.startswith(('2', '6'))]

        df_finished_parts['訂單'] = df_finished_parts['訂單'].astype(str)
        df_finished_parts = df_finished_parts[df_finished_parts['訂單'].str.startswith(('1', '6'))]
        
        # 因為欄位相同，對預備半品套用與撥料、成品撥料相同的訂單篩選邏輯
        df_prep_semi_finished['訂單'] = df_prep_semi_finished['訂單'].astype(str)
        df_prep_semi_finished = df_prep_semi_finished[df_prep_semi_finished['訂單'].str.startswith(('1', '2', '6'))]

        # DEBUG: 打印出原始 DataFrame 的欄位
        app_logger.info(f"DEBUG: df_wip_parts 欄位: {df_wip_parts.columns.tolist()}")
        app_logger.info(f"DEBUG: df_finished_parts 欄位: {df_finished_parts.columns.tolist()}")
        app_logger.info(f"DEBUG: df_inventory 欄位: {df_inventory.columns.tolist()}")

        df_specs = pd.read_excel(r'P:\F004\MPS維護\工單規格總表.xlsx')

        # 載入工單總表
        work_order_summary_path = Config.WORK_ORDER_BOOK_NAME  # 從設定檔讀取檔案名稱
        df_work_order_summary = pd.DataFrame() # 初始化為空 DataFrame
        if os.path.exists(work_order_summary_path):
            try:
                df_work_order_summary = pd.read_excel(work_order_summary_path, sheet_name='工單總表') # 更正頁籤名稱
                # 重新命名欄位以匹配預期
                if '品號說明' in df_work_order_summary.columns and '物料說明' not in df_work_order_summary.columns:
                    df_work_order_summary.rename(columns={'品號說明': '物料說明'}, inplace=True)
                app_logger.info(f"DEBUG: df_work_order_summary 欄位: {df_work_order_summary.columns.tolist()}")
            except Exception as e:
                app_logger.error(f"載入 '{Config.WORK_ORDER_BOOK_NAME}' 的 '工單總表' 頁籤時發生錯誤: {e}")
        else:
            app_logger.warning(f"警告：找不到 '{work_order_summary_path}' 檔案。工單摘要資訊將無法載入。 সন")

        on_order_path = r'P:\F004\MPS維護\已訂未交.XLSX'
        if os.path.exists(on_order_path):
            df_on_order = pd.read_excel(on_order_path)
        else:
            app_logger.warning("警告：找不到 '已訂未交.XLSX' 檔案。在途數量將全部視為 0。")
            df_on_order = pd.DataFrame(columns=['物料', '仍待交貨〈數量〉'])

        df_demand = pd.concat([df_wip_parts, df_finished_parts, df_prep_semi_finished], ignore_index=True)
        df_total_demand = df_demand.groupby('物料')['未結數量 (EINHEIT)'].sum().reset_index()
        df_total_demand.rename(columns={'未結數量 (EINHEIT)': 'total_demand'}, inplace=True)
        
        df_demand['需求日期'] = pd.to_datetime(df_demand['需求日期'], errors='coerce')
        demand_details_map = df_demand.groupby('物料').apply(
            lambda x: x[['訂單', '未結數量 (EINHEIT)', '需求日期']].to_dict('records'), include_groups=False
        ).to_dict()

        df_total_on_order = df_on_order.groupby('物料')['仍待交貨〈數量〉'].sum().reset_index()
        df_total_on_order.rename(columns={'仍待交貨〈數量〉': 'on_order_stock'}, inplace=True)

        # 以總需求為基礎，確保所有有需求的物料都被包含
        df_main = df_total_demand.copy()
        # 合併庫存資訊，只選擇需要的欄位
        df_main = pd.merge(df_main, df_inventory[['物料', '物料說明', '儲存地點', '基礎計量單位', '未限制', '在途和移轉', '品質檢驗中', '限制使用庫存', '閒置天數']], on='物料', how='left')
        # 合併在途數量
        df_main = pd.merge(df_main, df_total_on_order, on='物料', how='left')
        df_main['total_demand'] = df_main['total_demand'].fillna(0)
        df_main['on_order_stock'] = df_main['on_order_stock'].fillna(0)

        df_main.rename(columns={'未限制': 'unrestricted_stock', '品質檢驗中': 'inspection_stock'}, inplace=True)
        numeric_cols = ['unrestricted_stock', 'inspection_stock', 'total_demand', 'on_order_stock']
        for col in numeric_cols:
            df_main[col] = pd.to_numeric(df_main[col], errors='coerce').fillna(0)

        df_main['current_shortage'] = df_main['total_demand'] - (df_main['unrestricted_stock'] + df_main['inspection_stock'])
        df_main['projected_shortage'] = df_main['total_demand'] - (df_main['unrestricted_stock'] + df_main['inspection_stock'] + df_main['on_order_stock'])
        df_main['current_shortage'] = df_main['current_shortage'].clip(lower=0)
        df_main['projected_shortage'] = df_main['projected_shortage'].clip(lower=0)

        # 確保物料說明欄位不為空，優先從 df_demand 中獲取
        df_material_descriptions = df_demand[['物料', '物料說明']].drop_duplicates(subset=['物料'])
        df_main = pd.merge(df_main, df_material_descriptions, on='物料', how='left', suffixes=('', '_demand'))
        # 如果 df_main 中的物料說明為空，則使用來自 df_demand 的物料說明
        df_main['物料說明'] = df_main['物料說明'].fillna(df_main['物料說明_demand'])
        df_main.drop(columns=['物料說明_demand'], inplace=True)

        # 排除物料號碼以 '08' 開頭的物料
        df_main = df_main[~df_main['物料'].astype(str).str.startswith('08')]
        df_main['base_material_id'] = df_main['物料'].astype(str).str[:10]

        # --- 新增：建立訂單詳情對應表 ---
        df_order_materials = pd.concat([df_wip_parts, df_finished_parts], ignore_index=True)
        df_order_materials = df_order_materials[~df_order_materials['物料'].astype(str).str.startswith('08')] # 排除以 '08' 開頭的物料
        df_order_materials = pd.merge(df_order_materials, df_inventory, on='物料', how='left')
        df_order_materials.rename(columns={
            '未限制': 'unrestricted_stock',
            '品質檢驗中': 'inspection_stock'
        }, inplace=True)

        numeric_cols_order = ['未結數量 (EINHEIT)', 'unrestricted_stock', 'inspection_stock']
        for col in numeric_cols_order:
            df_order_materials[col] = pd.to_numeric(df_order_materials[col], errors='coerce').fillna(0)
        app_logger.info(f"DEBUG: df_order_materials (after numeric conversion) 的欄位為: {df_order_materials.columns.tolist()}")

        df_order_materials['order_shortage'] = (df_order_materials['未結數量 (EINHEIT)'] - (df_order_materials['unrestricted_stock'] + df_order_materials['inspection_stock'])).clip(lower=0)

        # 確保訂單號碼是字串，以便匹配
        order_details_map = df_order_materials.groupby('訂單').apply(
            lambda x: x[[ 
                '物料', '物料說明_x', '需求數量 (EINHEIT)', '領料數量 (EINHEIT)',
                '未結數量 (EINHEIT)', '需求日期', 'unrestricted_stock',
                'inspection_stock', 'order_shortage'
            ]].rename(columns={'物料說明_x': '物料說明'}).to_dict('records'), include_groups=False
        ).to_dict()

        # 確保規格表中的「訂單」欄位是字串，並只提取數字部分作為訂單號碼
        df_specs['訂單'] = df_specs['訂單'].astype(str).str.extract(r'(\d+)')[0]
        # 修正：在轉換為字典前，將所有 NaN 值替換為空字串，避免 JSON 序列化錯誤
        specs_map = df_specs.fillna('').groupby('訂單').apply(
            lambda x: x.to_dict('records'), include_groups=False
        ).to_dict()

        app_logger.info("資料載入與處理完畢。")
        app_logger.info("資料載入與處理完畢。")
        app_logger.info(f"DEBUG: order_details_map 中包含的訂單: {list(order_details_map.keys())[:5]}... (前5個)")
        app_logger.info(f"DEBUG: specs_map 中包含的訂單: {list(specs_map.keys())[:5]}... (前5個)")

        # 輔助函式：遞迴地將字典或列表中的 NaN 替換為空字串
        def replace_nan_in_dict(obj):
            if isinstance(obj, dict):
                return {k: replace_nan_in_dict(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [replace_nan_in_dict(elem) for elem in obj]
            elif isinstance(obj, float) and math.isnan(obj):
                return ''
            return obj

        # 在返回資料前，將所有 NaN 值替換為空字串，以確保 JSON 有效
        materials_dashboard_cleaned = df_main.fillna('').to_dict(orient='records')
        specs_data_cleaned = df_specs.fillna('').to_dict(orient='records')
        
        # 提取工單總表摘要資訊
        order_summary_map = {}
        if not df_work_order_summary.empty:
            required_cols = [
                '工單號碼', '下單客戶名稱', '物料說明', '生產開始', '生產結束',
                '機械外包', '電控外包', '噴漆外包', '鏟花外包', '捆包外包'
            ]
            # 確保所有需要的欄位都存在
            existing_cols = [col for col in required_cols if col in df_work_order_summary.columns]
            if len(existing_cols) != len(required_cols):
                app_logger.warning(f"警告：'{Config.WORK_ORDER_BOOK_NAME}' 中缺少部分預期欄位。預期: {required_cols}, 實際: {df_work_order_summary.columns.tolist()}")
            
            # 篩選出包含所有必要欄位的 DataFrame
            df_filtered_summary = df_work_order_summary[existing_cols].copy()
            
            # 處理日期欄位，轉換為字串格式
            for date_col in ['生產開始', '生產結束']:
                if date_col in df_filtered_summary.columns:
                    df_filtered_summary[date_col] = pd.to_datetime(df_filtered_summary[date_col], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')

            # 處理 NaN 值
            df_filtered_summary = df_filtered_summary.fillna('')

            # 處理重複的 '工單號碼'，保留第一個
            if '工單號碼' in df_filtered_summary.columns:
                df_filtered_summary['工單號碼'] = df_filtered_summary['工單號碼'].astype(str) # 確保工單號碼是字串
                df_filtered_summary.drop_duplicates(subset=['工單號碼'], keep='first', inplace=True)

            # 根據 '工單號碼' 建立映射
            order_summary_map = df_filtered_summary.set_index('工單號碼').to_dict(orient='index')
            app_logger.info(f"DEBUG: order_summary_map 中包含的工單: {list(order_summary_map.keys())[:5]}... (前5個)")


        # 對於巢狀字典，遞迴處理 NaN
        demand_details_map_cleaned = replace_nan_in_dict(demand_details_map)
        order_details_map_cleaned = replace_nan_in_dict(order_details_map)
        
        return {
            "materials_dashboard": materials_dashboard_cleaned,
            "specs_data": specs_data_cleaned,
            "demand_details_map": demand_details_map_cleaned,
            "order_details_map": order_details_map_cleaned,
            "specs_map": specs_map, # specs_map 已經在生成時處理過 NaN
            "order_summary_map": order_summary_map # 新增工單摘要資訊
        }

    except FileNotFoundError as e:
        app_logger.error(f"錯誤：找不到必要的資料檔案。請確認檔案是否存在且路徑正確：{e.filename}", exc_info=True)
        return None
    except Exception as e:
        app_logger.error(f"處理資料時發生未預期的錯誤: {e}", exc_info=True) # exc_info=True 會打印完整的堆疊追蹤
        return None

def update_cache_periodically():
    global live_cache_pointer
    while True:
        time.sleep(1800)
        app_logger.info("背景執行緒：準備更新快取...")
        
        # 在載入資料前先執行規格檔案的彙總
        app_logger.info("背景執行緒：執行工單規格檔案彙總...")
        try:
            consolidate_spec_files()
            app_logger.info("背景執行緒：工單規格檔案彙總完成。")
        except Exception as e:
            app_logger.error(f"背景執行緒：工單規格檔案彙總失敗: {e}", exc_info=True)
            # 如果彙總失敗，可以選擇是否繼續載入資料，這裡選擇繼續，但會記錄錯誤

        target_buffer = "B" if live_cache_pointer == "A" else "A"
        new_data = load_and_process_data()
        if new_data:
            data_cache[target_buffer] = new_data
            with cache_lock:
                live_cache_pointer = target_buffer
            app_logger.info(f"背景執行緒：快取更新完畢，線上服務已切換至緩衝區 {live_cache_pointer}")
        else:
            app_logger.error("背景執行緒：資料載入失敗，本次不更新快取。")

app = Flask(__name__)
app.secret_key = 'your_super_secret_key' # 在實際應用中，請使用更安全的金鑰並從環境變數中讀取

# 硬編碼用戶名和密碼 (僅用於示範)
USERS = {
    "admin": "password123"
}

# --- 路由 ---
@app.route('/')
def root():
    return redirect(url_for('order_query_page'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in USERS and USERS[username] == password:
            session['logged_in'] = True
            flash('登入成功！', 'success')
            return redirect(url_for('admin_dashboard_page'))
        else:
            flash('無效的用戶名或密碼', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash('您已登出', 'info')
    return redirect(url_for('login'))

@app.route('/procurement')
def procurement_page():
    record_page_view('procurement.html', request.remote_addr)
    return render_template('procurement.html')

@app.route('/order_query')
def order_query_page():
    record_page_view('order_query.html', request.remote_addr)
    return render_template('order_query.html')

# --- API 端點 ---
@app.route('/api/materials')
def get_materials():
    with cache_lock:
        current_data = data_cache[live_cache_pointer]
    if current_data and "materials_dashboard" in current_data:
        return jsonify(current_data["materials_dashboard"])
    return jsonify([])

@app.route('/api/material/<material_id>/details')
def get_material_details(material_id):
    try:
        with cache_lock:
            current_data = data_cache[live_cache_pointer]
        
        if not current_data:
            app_logger.error("get_material_details: 資料尚未載入")
            return jsonify({"error": "資料尚未載入"}), 500

        all_materials_df = pd.DataFrame(current_data.get("materials_dashboard", []))
        if all_materials_df.empty:
            app_logger.error("get_material_details: 物料資料庫為空")
            return jsonify({"error": "物料資料庫為空"}), 500

        # 1. 獲取庫存總覽
        material_info = all_materials_df[all_materials_df['物料'].astype(str) == material_id]
        if material_info.empty:
            app_logger.warning(f"get_material_details: 找不到物料 {material_id}")
            return jsonify({"error": "找不到該物料"}), 404
        
        stock_summary = material_info.iloc[0].to_dict()
        total_available_stock = stock_summary.get('unrestricted_stock', 0) + stock_summary.get('inspection_stock', 0)

        # 2. 獲取、過濾、排序需求詳情
        demand_map = current_data.get("demand_details_map", {})
        demand_details = [d.copy() for d in demand_map.get(material_id, [])]
        demand_details = [d for d in demand_details if d.get('未結數量 (EINHEIT)', 0) > 0]

        demand_details.sort(key=lambda x: x.get('需求日期') or pd.Timestamp.max, reverse=False)

        # 3. 執行庫存消耗計算
        running_stock = total_available_stock
        shortage_triggered = False
        for item in demand_details:
            demand_qty = item.get('未結數量 (EINHEIT)', 0)
            running_stock -= demand_qty
            item['remaining_stock'] = running_stock
            if running_stock < 0 and not shortage_triggered:
                shortage_triggered = True
            item['is_shortage_point'] = shortage_triggered
            if pd.notna(item['需求日期']):
                item['需求日期'] = item['需求日期'].strftime('%Y-%m-%d')

        # 4. 獲取替代品庫存
        base_id = str(material_id)[:10]
        substitutes_df = all_materials_df[
            (all_materials_df['base_material_id'] == base_id) & 
            (all_materials_df['物料'].astype(str) != material_id)
        ]
        substitute_inventory = substitutes_df[['物料', '物料說明', 'unrestricted_stock', 'inspection_stock']].to_dict('records')

        return jsonify({
            "stock_summary": {
                "unrestricted": stock_summary.get('unrestricted_stock', 0),
                "inspection": stock_summary.get('inspection_stock', 0),
                "on_order": stock_summary.get('on_order_stock', 0)
            },
            "demand_details": demand_details,
            "substitute_inventory": substitute_inventory
        })

    except Exception as e:
        app_logger.error(f"在 get_material_details 函式中發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "一個後端錯誤發生了"}), 500

@app.route('/api/order/<order_id>')
def get_order_details(order_id):
    try:
        with cache_lock:
            current_data = data_cache[live_cache_pointer]
        
        if not current_data:
            app_logger.error("get_order_details: 資料尚未載入")
            return jsonify({"error": "資料尚未載入"}), 500

        # 從快取中獲取訂單的備註和版本
        with order_note_cache_lock:
            cached_info = order_note_cache.get(str(order_id).strip(), {})
        order_note = cached_info.get('note')
        spec_version = cached_info.get('version')

        order_materials = current_data.get("order_details_map", {}).get(order_id, [])
        raw_order_specs = current_data.get("specs_map", {}).get(order_id, [])
        order_summary = current_data.get("order_summary_map", {}).get(order_id, {})

        filtered_order_specs = _filter_order_specs(raw_order_specs)

        for item in order_materials:
            if isinstance(item.get('需求日期'), pd.Timestamp):
                item['需求日期'] = item['需求日期'].strftime('%Y-%m-%d')
            elif pd.isna(item.get('需求日期')):
                item['需求日期'] = ''

        return jsonify({
            "order_summary": order_summary,
            "order_note": order_note,
            "spec_version": spec_version, # 新增訂單版本到 API 回應
            "order_specs": filtered_order_specs,
            "order_materials": order_materials
        })

    except Exception as e:
        app_logger.error(f"在 get_order_details 函式中發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "一個後端錯誤發生了"}), 500

def _filter_order_specs(raw_specs):
    """
    輔助函式：篩選並重新排序訂單規格的欄位，使其與前端顯示一致。
    """
    filtered_specs = []
    for spec_item in raw_specs:
        filtered_spec = {
            '內部特性號碼': spec_item.get('內部特性號碼', ''),
            '特性說明': spec_item.get('特性說明', ''),
            '特性值': spec_item.get('特性值', ''),
            '值說明': spec_item.get('值說明', '')
        }
        filtered_specs.append(filtered_spec)
    return filtered_specs

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

@app.route('/api/download_specs/<order_id>')
def download_specs(order_id):
    try:
        with cache_lock:
            current_data = data_cache[live_cache_pointer]
        
        if not current_data:
            app_logger.error("download_specs: 資料尚未載入")
            return jsonify({"error": "資料尚未載入"}), 500

        specs_map = current_data.get("specs_map", {})
        raw_order_specs = specs_map.get(order_id, [])

        if not raw_order_specs:
            app_logger.warning(f"download_specs: 找不到訂單 {order_id} 的規格資料")
            return jsonify({"error": f"找不到訂單 {order_id} 的規格資料"}), 404

        # 使用輔助函式篩選規格數據
        filtered_order_specs_for_download = _filter_order_specs(raw_order_specs)
        
        df_order_specs = pd.DataFrame(filtered_order_specs_for_download)
        
        # 將 DataFrame 寫入記憶體中的 Excel 檔案
        output = io.BytesIO()
        df_order_specs.to_excel(output, index=False, sheet_name=f'訂單_{order_id}_規格')
        output.seek(0)

        # 使用 openpyxl 進行後續處理
        wb = load_workbook(output)
        ws = wb.active

        # 1. 設定格線
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # 2. 自動調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column) # 獲取欄位字母
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # 調整寬度，增加一些緩衝
            ws.column_dimensions[column_letter].width = adjusted_width

        # 將修改後的 workbook 存回 BytesIO
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # --- 修正下載問題 ---
        # 建立檔名
        filename = f'訂單_{order_id}_規格.xlsx'
        
        # 建立回應
        response = make_response(final_output.getvalue())
        
        # 設定標頭
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = (
            f"attachment; filename={quote(filename.encode('utf-8'))}; "
            f"filename*=UTF-8''{quote(filename.encode('utf-8'))}"
        )
        
        return response

    except Exception as e:
        app_logger.error(f"下載訂單 {order_id} 規格時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "無法下載檔案"}), 500

@app.route('/admin_dashboard')
def admin_dashboard_page():
    if not session.get('logged_in'):
        flash('請先登入才能訪問此頁面', 'warning')
        return redirect(url_for('login'))
    record_page_view('admin_dashboard.html', request.remote_addr)
    return render_template('admin_dashboard.html')

@app.route('/api/admin/traffic')
def get_traffic_data():
    views_data = read_views()
    traffic_summary = {}
    
    for page, data in views_data.items():
        total_views = data.get("total_views", 0)
        ip_access_times = data.get("ip_access_times", {})
        
        ip_stats = []
        for ip, timestamps in ip_access_times.items():
            visits = len(timestamps)
            last_visit = timestamps[-1] if timestamps else None
            ip_stats.append({
                "ip": ip,
                "visits": visits,
                "last_visit": last_visit
            })
        
        traffic_summary[page] = {
            "total_views": total_views,
            "ip_stats": ip_stats
        }
        
    return jsonify(traffic_summary)

@app.route('/api/status')
def api_status():
    with cache_lock:
        current_data = data_cache[live_cache_pointer]
    status = {
        "service_status": "online",
        "live_cache": live_cache_pointer,
        "data_loaded": current_data is not None
    }
    return jsonify(status)

if __name__ == '__main__':
    app_logger.info("主程式：執行首次工單規格檔案彙總...")
    try:
        consolidate_spec_files()
        app_logger.info("主程式：首次工單規格檔案彙總完成。")
    except Exception as e:
        app_logger.error(f"主程式：首次工單規格檔案彙總失敗: {e}", exc_info=True)

    app_logger.info("主程式：執行首次資料載入...")
    initial_data = load_and_process_data()
    if initial_data:
        data_cache[live_cache_pointer] = initial_data
        app_logger.info("主程式：首次資料載入成功。")
    else:
        app_logger.error("主程式：首次資料載入失敗！服務將在沒有資料的情況下啟動。")

    app_logger.info("主程式：執行首次訂單備註與版本快取載入...")
    load_order_notes_to_cache()

    cache_updater_thread = threading.Thread(target=update_cache_periodically, daemon=True)
    cache_updater_thread.start()

    order_note_cache_updater_thread = threading.Thread(target=update_order_note_cache_periodically, daemon=True)
    order_note_cache_updater_thread.start()

    serve(app, host='0.0.0.0', port=5002, threads=4)
