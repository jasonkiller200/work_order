# app/services/spec_service.py
# 規格檔案服務

import os
import pandas as pd
import io
import logging
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from app.config import FilePaths

app_logger = logging.getLogger(__name__)

class SpecService:
    """規格檔案服務"""
    
    @staticmethod
    def consolidate_spec_files():
        """
        將指定資料夾內的所有 Excel 規格檔案合併成一個總表
        """
        source_folder = FilePaths.SPEC_SOURCE_FOLDER
        output_filename = FilePaths.SPEC_OUTPUT_FILE
        
        # 檢查來源資料夾是否存在
        if not os.path.isdir(source_folder):
            app_logger.error(f"錯誤：找不到名為 '{source_folder}' 的資料夾。")
            return
        
        all_specs_data = []
        app_logger.info(f"開始掃描 '{source_folder}' 資料夾...")
        
        # 遍歷、讀取、處理檔案
        for filename in os.listdir(source_folder):
            if filename.lower().endswith(('.xlsx', '.xls')):
                file_path = os.path.join(source_folder, filename)
                
                try:
                    # 從檔名提取訂單號碼（去除副檔名，並只取前9個字元）
                    order_number = os.path.splitext(filename)[0][:9]
                    
                    # 讀取 Excel 檔案
                    df = pd.read_excel(file_path)
                    
                    # 在第一欄插入訂單號碼
                    df.insert(0, '訂單', order_number)
                    
                    all_specs_data.append(df)
                    app_logger.info(f"  - 已處理檔案: {filename} (訂單: {order_number})")
                
                except Exception as e:
                    app_logger.error(f"  - 處理檔案 {filename} 時發生錯誤: {e}")
        
        # 合併與儲存
        if not all_specs_data:
            app_logger.warning("在資料夾中沒有找到任何 Excel 檔案可以處理。")
            return
        
        # 將所有 DataFrame 合併成一個
        consolidated_df = pd.concat(all_specs_data, ignore_index=True)
        
        # 確保輸出目錄存在
        output_dir = os.path.dirname(output_filename)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            app_logger.info(f"已建立輸出目錄: {output_dir}")
        
        # 將合併後的 DataFrame 寫入新的 Excel 檔案
        consolidated_df.to_excel(output_filename, index=False)
        app_logger.info(f"成功！已將 {len(all_specs_data)} 個規格檔案合併至 '{output_filename}'。")
    
    @staticmethod
    def filter_order_specs(raw_specs):
        """
        篩選並重新排序訂單規格的欄位
        
        Args:
            raw_specs: 原始規格資料列表
            
        Returns:
            篩選後的規格資料列表
        """
        filtered_specs = []
        for spec_item in raw_specs:
            filtered_spec = {
                '內部特性號碼': spec_item.get('內部特性號碼', ''),
                '特性說明': spec_item.get('特性說明', ''),
                '特性值': spec_item.get('特性值', ''),
                '值說明': spec_item.get('值說明', '')
            }
            filtered_specs.append(filtered_spec)
        return filtered_specs
    
    @staticmethod
    def generate_spec_excel(order_id, raw_order_specs):
        """
        產生規格 Excel 檔案
        
        Args:
            order_id: 訂單 ID
            raw_order_specs: 原始規格資料
            
        Returns:
            Excel 檔案的二進位內容
        """
        # 使用輔助函式篩選規格數據
        filtered_order_specs = SpecService.filter_order_specs(raw_order_specs)
        
        df_order_specs = pd.DataFrame(filtered_order_specs)
        
        # 將 DataFrame 寫入記憶體中的 Excel 檔案
        output = io.BytesIO()
        df_order_specs.to_excel(output, index=False, sheet_name=f'訂單_{order_id}_規格')
        output.seek(0)
        
        # 使用 openpyxl 進行後續處理
        wb = load_workbook(output)
        ws = wb.active
        
        # 1. 設定格線
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # 2. 自動調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 將修改後的 workbook 存回 BytesIO
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        return final_output.getvalue()
