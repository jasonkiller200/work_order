"""
Excel 同步服務 - 將系統交期資料同步至外部 Excel 檔案
"""
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import func
from ..models.database import db, DeliverySchedule

app_logger = logging.getLogger(__name__)


# Excel 檔案設定
EXCEL_CONFIG = {
    'file_path': r'Q:\G003\生產排程\8週生產排程紀錄\裝配進度&缺料情報\第一廠\組件課\未來半品缺料.xlsm',
    'sheet_name': '半品',
    'material_col': 'A',  # 料號欄位
    'status_col': 'G',    # 狀態欄位
    'delivery_col': 'H',  # 交期欄位
    'header_row': 1       # 表頭列數
}


def _parse_excel_date(val):
    """將 Excel 讀取出的多種交期格式安全地轉為 datetime.date"""
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date'):  # 已是 date 物件
        return val
    if isinstance(val, str):
        val = val.strip()
        # 嘗試解析常見格式
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def sync_delivery_to_excel():
    """
    將採購儀表板的第一筆預計交貨日期同步到 Excel 檔案
    
    Returns:
        dict: 同步結果統計
            - success: bool
            - synced_count: int - 成功同步筆數
            - cleared_count: int - 清空筆數（過期且狀態為 OK/缺料）
            - skipped_count: int - 跳過筆數（系統無交期）
            - error: str - 錯誤訊息（如果有）
    """
    result = {
        'success': False,
        'synced_count': 0,
        'cleared_count': 0,
        'skipped_count': 0,
        'not_found_count': 0,
        'error': None
    }
    
    try:
        file_path = EXCEL_CONFIG['file_path']
        
        # 1. 檢查檔案是否存在
        if not os.path.exists(file_path):
            result['error'] = f'Excel 檔案不存在: {file_path}'
            app_logger.error(result['error'])
            return result
        
        # 2. 載入 Excel 檔案 (保留巨集)
        app_logger.info(f'正在載入 Excel 檔案: {file_path}')
        wb = load_workbook(file_path, keep_vba=True)
        
        # 3. 取得指定頁籤
        sheet_name = EXCEL_CONFIG['sheet_name']
        if sheet_name not in wb.sheetnames:
            result['error'] = f'找不到頁籤: {sheet_name}'
            app_logger.error(result['error'])
            return result
        
        ws = wb[sheet_name]
        
        # 4. 查詢系統中所有有交期的物料 (取每個物料的第一筆有效交期)
        # 使用子查詢取得每個物料最近的交期
        delivery_map = _get_first_delivery_dates()
        app_logger.info(f'系統中共有 {len(delivery_map)} 個物料有交期資料')
        
        # 5. 遍歷 Excel 每一列
        header_row = EXCEL_CONFIG['header_row']
        material_col = EXCEL_CONFIG['material_col']
        status_col = EXCEL_CONFIG['status_col']
        delivery_col = EXCEL_CONFIG['delivery_col']
        
        from ..utils.helpers import get_taiwan_time
        today = get_taiwan_time().date()
        
        for row_num in range(header_row + 1, ws.max_row + 1):
            # 取得料號
            material_id = ws[f'{material_col}{row_num}'].value
            
            if not material_id:
                continue
            
            # 轉為字串並去除空白
            material_id = str(material_id).strip()
            
            # 取得 G 欄狀態與 H 欄原有交期
            status_val = ws[f'{status_col}{row_num}'].value
            status_str = str(status_val).strip() if status_val is not None else ""
            
            current_delivery_val = ws[f'{delivery_col}{row_num}'].value
            
            # 取得系統對應的交期
            system_date = delivery_map.get(material_id)  # datetime.date 物件或 None
            
            # 決定要評估的日期：優先採用系統新交期，其次為 Excel 原有交期
            eval_date = system_date if system_date else _parse_excel_date(current_delivery_val)
            
            # 檢查是否符合清理條件：交期已小於今天，且 G 欄為 "OK" 或 "缺料"
            is_expired = eval_date is not None and eval_date < today
            is_target_status = status_str in ("OK", "缺料")
            
            if is_expired and is_target_status:
                # 符合清理條件，將 H 欄清空
                ws[f'{delivery_col}{row_num}'] = None
                result['cleared_count'] += 1
            elif system_date:
                # 不符合清理條件，且系統中有最新交期，回填新交期
                ws[f'{delivery_col}{row_num}'] = system_date.strftime('%Y-%m-%d')
                result['synced_count'] += 1
            else:
                # 系統無交期且不符合清理條件，保留 Excel 原有值
                result['skipped_count'] += 1
        
        # 7. 儲存 Excel
        app_logger.info(f'正在儲存 Excel 檔案...')
        wb.save(file_path)
        wb.close()
        
        result['success'] = True
        app_logger.info(
            f'交期同步完成: 成功 {result["synced_count"]} 筆, '
            f'清空 {result["cleared_count"]} 筆, '
            f'跳過 {result["skipped_count"]} 筆'
        )
        
    except PermissionError:
        result['error'] = 'Excel 檔案正在被其他程式使用中，請先關閉檔案後再試'
        app_logger.error(result['error'])
    except Exception as e:
        result['error'] = str(e)
        app_logger.error(f'交期同步失敗: {e}', exc_info=True)
    
    return result


def _get_first_delivery_dates():
    """
    取得每個物料的第一筆有效交期
    
    Returns:
        dict: {material_id: expected_date}
    """
    # 使用 SQL 子查詢取得每個物料的最近交期
    subquery = db.session.query(
        DeliverySchedule.material_id,
        func.min(DeliverySchedule.expected_date).label('first_date')
    ).filter(
        DeliverySchedule.status.notin_(['cancelled', 'completed']),
        DeliverySchedule.expected_date.isnot(None)
    ).group_by(
        DeliverySchedule.material_id
    ).subquery()
    
    # 取得完整的交期記錄
    results = db.session.query(
        subquery.c.material_id,
        subquery.c.first_date
    ).all()
    
    return {row.material_id: row.first_date for row in results}
